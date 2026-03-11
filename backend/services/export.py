"""
Service Export — Génération PDF et Excel professionnels.
"""
from __future__ import annotations

import io
from datetime import date
from typing import Dict, List, Optional

import pandas as pd


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def export_dashboard_excel(matieres: List[dict], dept: str) -> bytes:
    """
    Génère un fichier Excel multi-feuilles :
    - Feuille "Résumé" : KPIs globaux
    - Feuille par classe
    - Feuille "Alertes"
    """
    output = io.BytesIO()

    df_all = pd.DataFrame(matieres)
    if df_all.empty:
        df_all = pd.DataFrame(columns=["matiere", "classe", "vhp", "vhr", "taux", "ecart", "statut"])

    # Renommer pour lisibilité
    rename = {
        "matiere": "Matière", "classe": "Classe", "responsable": "Responsable",
        "email": "Email", "semestre": "Semestre",
        "vhp": "VHP (h)", "vhr": "VHR (h)", "taux": "Taux",
        "ecart": "Écart (h)", "statut": "Statut",
        "proba_retard": "Risque retard", "prediction_label": "Prédiction",
    }
    df_all = df_all.rename(columns={k: v for k, v in rename.items() if k in df_all.columns})

    if "Taux" in df_all.columns:
        df_all["Taux (%)"] = (df_all["Taux"] * 100).round(1)
        df_all = df_all.drop(columns=["Taux"])

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        # ── Feuille Résumé KPIs ──────────────────────────────────────────
        kpi_data = {
            "Indicateur": [
                "Département", "Date export", "Total matières",
                "VHP total (h)", "VHR total (h)", "Taux global (%)",
                "Terminées", "En cours", "Non démarrées",
            ],
            "Valeur": [
                dept,
                date.today().strftime("%d/%m/%Y"),
                len(df_all),
                df_all["VHP (h)"].sum() if "VHP (h)" in df_all.columns else 0,
                df_all["VHR (h)"].sum() if "VHR (h)" in df_all.columns else 0,
                f"{df_all['Taux (%)'].mean():.1f}" if "Taux (%)" in df_all.columns else "0",
                (df_all["Statut"] == "Terminé").sum() if "Statut" in df_all.columns else 0,
                (df_all["Statut"] == "En cours").sum() if "Statut" in df_all.columns else 0,
                (df_all["Statut"] == "Non démarré").sum() if "Statut" in df_all.columns else 0,
            ],
        }
        pd.DataFrame(kpi_data).to_excel(writer, sheet_name="Résumé KPIs", index=False)

        # ── Feuille globale ───────────────────────────────────────────────
        cols_export = [c for c in ["Classe", "Matière", "Responsable", "Semestre",
                                    "VHP (h)", "VHR (h)", "Taux (%)", "Écart (h)", "Statut",
                                    "Prédiction"] if c in df_all.columns]
        df_all[cols_export].sort_values(["Classe", "Taux (%)"]).to_excel(
            writer, sheet_name="Tous enseignements", index=False
        )

        # ── Feuilles par classe ───────────────────────────────────────────
        if "Classe" in df_all.columns:
            for classe in sorted(df_all["Classe"].unique())[:10]:  # Max 10 onglets
                df_cls = df_all[df_all["Classe"] == classe][cols_export].copy()
                sheet_name = str(classe)[:31]
                df_cls.to_excel(writer, sheet_name=sheet_name, index=False)

        # ── Feuille alertes ───────────────────────────────────────────────
        alertes_cols = [c for c in ["Classe", "Matière", "Responsable", "Email",
                                     "VHP (h)", "VHR (h)", "Écart (h)", "Statut"] if c in df_all.columns]
        if "Écart (h)" in df_all.columns:
            df_alertes = df_all[df_all["Écart (h)"] < -6][alertes_cols].sort_values("Écart (h)")
            df_alertes.to_excel(writer, sheet_name="Alertes", index=False)

        # ── Formatage ─────────────────────────────────────────────────────
        _format_workbook(writer.book)

    return output.getvalue()


def _format_workbook(wb):
    """Applique le formatage professionnel à toutes les feuilles."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="0B3D91")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    ALT_FILL   = PatternFill("solid", fgColor="F0F4FF")
    BORDER     = Border(
        bottom=Side(style="thin", color="D0D8F0"),
        right=Side(style="thin", color="D0D8F0"),
    )

    for sheet in wb.worksheets:
        # En-têtes
        for cell in sheet[1]:
            cell.fill  = HEADER_FILL
            cell.font  = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Lignes alternées + bordures
        for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.border = BORDER
                if i % 2 == 0:
                    cell.fill = ALT_FILL

        # Largeur colonnes auto
        for col_idx in range(1, sheet.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                (len(str(sheet.cell(row=r, column=col_idx).value or ""))
                 for r in range(1, min(sheet.max_row + 1, 50))),
                default=8,
            )
            sheet.column_dimensions[col_letter].width = min(max_len + 4, 45)

        # Figer la première ligne
        sheet.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT PDF
# ─────────────────────────────────────────────────────────────────────────────

def export_dashboard_pdf(
    matieres: List[dict],
    dept: str,
    dept_config: Optional[dict] = None,
    logo_path: Optional[str] = None,
) -> bytes:
    """Génère un PDF professionnel du dashboard."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm, mm
        from reportlab.platypus import (
            Image, Paragraph, SimpleDocTemplate, Spacer,
            Table, TableStyle
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        return b"ReportLab non disponible"

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"],
        fontSize=16, textColor=colors.HexColor("#0B3D91"), spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "Sub", parent=styles["Normal"],
        fontSize=10, textColor=colors.grey, spaceAfter=12,
    )
    cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=8)

    story = []

    # En-tête
    dept_long = (dept_config or {}).get("department_long", dept)
    story.append(Paragraph(f"Dashboard Suivi des Enseignements — {dept_long}", title_style))
    story.append(Paragraph(f"Exporté le {date.today().strftime('%d/%m/%Y')}", subtitle_style))
    story.append(Spacer(1, 8*mm))

    if not matieres:
        story.append(Paragraph("Aucune donnée disponible.", styles["Normal"]))
        doc.build(story)
        return buffer.getvalue()

    df = pd.DataFrame(matieres)

    # KPIs résumé
    total_vhp = df["vhp"].sum() if "vhp" in df else 0
    total_vhr = df["vhr"].sum() if "vhr" in df else 0
    taux_g = (total_vhr / total_vhp * 100) if total_vhp > 0 else 0
    nb_t = (df["statut"] == "Terminé").sum() if "statut" in df else 0
    nb_nd = (df["statut"] == "Non démarré").sum() if "statut" in df else 0

    kpi_table_data = [
        ["Matières", "VHP total", "VHR total", "Taux global", "Terminées", "Non démarrées"],
        [str(len(df)), f"{total_vhp:.0f}h", f"{total_vhr:.0f}h",
         f"{taux_g:.1f}%", str(nb_t), str(nb_nd)],
    ]
    kpi_tbl = Table(kpi_table_data, colWidths=[3.5*cm]*6)
    kpi_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B3D91")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F0F4FF"), colors.white]),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#D0D8F0")),
        ("FONTNAME",   (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 1), (-1, 1), 12),
    ]))
    story.append(kpi_tbl)
    story.append(Spacer(1, 8*mm))

    # Table principale (top 40 par écart)
    cols_pdf = ["classe", "matiere", "responsable", "vhp", "vhr", "ecart", "statut"]
    cols_labels = ["Classe", "Matière", "Responsable", "VHP", "VHR", "Écart", "Statut"]
    col_widths = [3*cm, 7*cm, 5*cm, 1.8*cm, 1.8*cm, 1.8*cm, 2.5*cm]

    header = [Paragraph(f"<b>{h}</b>", cell_style) for h in cols_labels]
    rows_data = [header]

    df_sorted = df.sort_values("ecart") if "ecart" in df.columns else df
    for _, row in df_sorted.head(40).iterrows():
        statut = str(row.get("statut", ""))
        ecart_val = float(row.get("ecart", 0))
        ecart_color = "#D93025" if ecart_val < -6 else "#0F172A"
        row_data = [
            Paragraph(str(row.get("classe", ""))[:20], cell_style),
            Paragraph(str(row.get("matiere", ""))[:60], cell_style),
            Paragraph(str(row.get("responsable", ""))[:30], cell_style),
            Paragraph(f"{float(row.get('vhp', 0)):.0f}h", cell_style),
            Paragraph(f"{float(row.get('vhr', 0)):.0f}h", cell_style),
            Paragraph(f'<font color="{ecart_color}">{ecart_val:.0f}h</font>', cell_style),
            Paragraph(statut, cell_style),
        ]
        rows_data.append(row_data)

    main_tbl = Table(rows_data, colWidths=col_widths, repeatRows=1)
    main_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B3D91")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFF")]),
        ("GRID",       (0, 0), (-1, -1), 0.3, colors.HexColor("#D0D8F0")),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(main_tbl)

    doc.build(story)
    return buffer.getvalue()
